"""Jira REST API 데이터 수집 모듈"""

import logging
import requests
import pandas as pd
from typing import Optional

from config import (
    JIRA_BASE_URL,
    JIRA_USERNAME,
    JIRA_PASSWORD,
    JIRA_ISSUE_TYPES,
    JIRA_MAX_RESULTS,
    CUSTOM_FIELD_MAP,
)

logger = logging.getLogger(__name__)


class JiraClient:
    """Jira REST API를 통해 SR 데이터를 수집하는 클라이언트"""

    def __init__(
        self,
        base_url: str = JIRA_BASE_URL,
        username: str = JIRA_USERNAME,
        password: str = JIRA_PASSWORD,
    ):
        self.base_url = base_url.rstrip("/")
        self.session = requests.Session()
        self.session.auth = (username, password)
        self.session.headers.update({"Content-Type": "application/json"})

    def _build_jql(self, project: str, from_date: str) -> str:
        """JQL 쿼리 생성"""
        issue_types = ", ".join(JIRA_ISSUE_TYPES)
        return (
            f'project in ({project}) '
            f'AND issuetype in ({issue_types}) '
            f'AND createdDate >= "{from_date}" '
            f'ORDER BY created DESC'
        )

    def _search(self, jql: str, start_at: int = 0) -> dict:
        """Jira REST API search 호출"""
        url = f"{self.base_url}/rest/api/2/search"
        params = {
            "jql": jql,
            "startAt": start_at,
            "maxResults": JIRA_MAX_RESULTS,
        }
        response = self.session.get(url, params=params, timeout=60, verify=False)
        response.raise_for_status()
        return response.json()

    def _extract_field_value(self, fields: dict, field_key: str):
        """Jira 필드 값을 추출 (중첩 객체 처리)"""
        value = fields.get(field_key)
        if value is None:
            return None
        if isinstance(value, dict):
            # name, displayName, value 순서로 탐색
            return (
                value.get("name")
                or value.get("displayName")
                or value.get("value")
                or str(value)
            )
        return value

    def _parse_issue(self, issue: dict) -> dict:
        """단일 이슈를 딕셔너리로 변환"""
        fields = issue["fields"]
        row = {
            "프로젝트": self._extract_field_value(fields, "project"),
            "키": issue["key"],
            "요약": fields.get("summary"),
            "이슈 유형": self._extract_field_value(fields, "issuetype"),
            "상태": self._extract_field_value(fields, "status"),
            "해결책": self._extract_field_value(fields, "resolution"),
            "담당자": self._extract_field_value(fields, "assignee"),
            "생성일": fields.get("created"),
            "변경일": fields.get("updated"),
            "해결일": fields.get("resolutiondate"),
            "설명": fields.get("description"),
        }
        # 커스텀 필드 매핑
        for cf_key, col_name in CUSTOM_FIELD_MAP.items():
            row[col_name] = self._extract_field_value(fields, cf_key)
        return row

    def fetch_issues(self, project: str, from_date: str) -> pd.DataFrame:
        """
        Jira에서 이슈 데이터를 수집하여 DataFrame으로 반환

        Args:
            project: Jira 프로젝트 코드 (예: "AMDP1")
            from_date: 시작일 (예: "2025-01-01")

        Returns:
            pd.DataFrame: 수집된 이슈 데이터
        """
        jql = self._build_jql(project, from_date)
        logger.info(f"JQL: {jql}")

        all_issues = []
        start_at = 0

        while True:
            logger.info(f"  페이지 조회 중... (startAt={start_at})")
            result = self._search(jql, start_at)

            issues = result.get("issues", [])
            if not issues:
                break

            for issue in issues:
                all_issues.append(self._parse_issue(issue))

            total = result.get("total", 0)
            start_at += len(issues)

            logger.info(f"  {start_at}/{total} 건 수집 완료")

            if start_at >= total:
                break

        logger.info(f"총 {len(all_issues)}건 수집 완료")
        return pd.DataFrame(all_issues)

    def fetch_custom_field_ids(self) -> dict:
        """
        Jira 커스텀 필드 ID와 이름 목록을 조회
        (설정 시 커스텀 필드 ID 확인용)
        """
        url = f"{self.base_url}/rest/api/2/field"
        response = self.session.get(url, timeout=30, verify=False)
        response.raise_for_status()
        fields = response.json()
        return {
            f["id"]: f["name"]
            for f in fields
            if f["id"].startswith("customfield_")
        }


def load_from_excel(file_path: str) -> pd.DataFrame:
    """
    Excel 파일에서 RAW 데이터를 로드 (Jira 연동 대신 사용)

    Args:
        file_path: Excel 파일 경로

    Returns:
        pd.DataFrame: 로드된 데이터
    """
    import openpyxl

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb["general_report"]

    # 헤더 읽기 (중복 컬럼명 처리)
    raw_headers = [cell.value for cell in ws[1]]
    headers = []
    seen = {}
    for h in raw_headers:
        if h in seen:
            seen[h] += 1
            headers.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 0
            headers.append(h)

    # 데이터 읽기
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append(list(row))

    wb.close()

    df = pd.DataFrame(data, columns=headers)

    # 중복 컬럼 정리: 두 번째 상태 컬럼을 실제 상태로 사용
    if "상태_1" in df.columns:
        df["상태"] = df["상태_1"]
        df.drop(columns=["상태_1"], inplace=True, errors="ignore")
    if "해결책_1" in df.columns:
        df.drop(columns=["해결책_1"], inplace=True, errors="ignore")
    if "시스템명_1" in df.columns:
        df.drop(columns=["시스템명_1"], inplace=True, errors="ignore")

    # 기존 수식 기반 컬럼 제거 (Python에서 재계산)
    for col in ["처리상태", "문의유형", "평균처리시간\n(생성일-접수일)", "평균처리시간\n(생성일-해결일)"]:
        if col in df.columns:
            df.drop(columns=[col], inplace=True, errors="ignore")

    logger.info(f"Excel에서 {len(df)}건 로드 완료 ({file_path})")
    return df
