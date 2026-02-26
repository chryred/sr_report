"""Excel 리포트 생성 모듈 (서식 포함)"""

import logging
import os
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from config import (
    INQUIRY_TYPE_ORDER,
    PROCESS_STATUS_ORDER,
    OUTPUT_DIR,
    PROJECT_TEAM_NAME_MAP,
    PROJECT_SHEET_PREFIX_MAP,
)
from data_processor import generate_summary_stats

logger = logging.getLogger(__name__)

# ── 스타일 정의 ─────────────────────────────────────────────
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

HEADER_FILL_DARK = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FILL_LIGHT = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
HEADER_FILL_YELLOW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
HEADER_FILL_GRAY = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

FONT_HEADER_WHITE = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
FONT_HEADER_BLACK = Font(name="맑은 고딕", size=10, bold=True)
FONT_NORMAL = Font(name="맑은 고딕", size=10)
FONT_TITLE = Font(name="맑은 고딕", size=12, bold=True)
FONT_SECTION = Font(name="맑은 고딕", size=11, bold=True)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

# 문의유형 짧은 이름 (리포트 헤더용)
INQUIRY_SHORT_NAMES = {
    "1. 기능문의": "기능문의",
    "2. 단순조치/운영지원": "단순조치/\n운영지원",
    "3. 데이터 추출 및 수정": "데이터\n추출/수정",
    "4. 신규개발 및 개선": "신규\n개발/개선",
    "5. 운영관리(시스템/품질/보안)": "운영관리\n(시스템/품질/보안)",
}


def _apply_cell_style(cell, font=None, fill=None, alignment=None, border=None):
    """셀에 스타일 적용"""
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border


def _set_column_widths(ws, widths: dict):
    """컬럼 너비 설정"""
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def generate_report(
    df: pd.DataFrame,
    stats: dict,
    project_name: str,
    report_month: str,
    output_path: str = None,
    selected_projects: list = None,
) -> str:
    """
    최종 Excel 리포트 생성

    Args:
        df: 정제된 DataFrame (selected_projects로 이미 필터된 상태)
        stats: generate_summary_stats()의 결과 (df 전체 기준)
        project_name: 프로젝트/팀 이름 (헤더 표시용)
        report_month: 리포트 대상 월 (예: "2025-02")
        output_path: 출력 파일 경로 (None이면 자동 생성)
        selected_projects: 선택된 프로젝트 목록. 지정 시 프로젝트별 시트 세트 생성.
                           None 또는 빈 리스트이면 df 전체 기준 단일 시트 세트 생성.

    Returns:
        str: 생성된 파일 경로
    """
    if output_path is None:
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(
            OUTPUT_DIR, f"SR_Report_{report_month}_{timestamp}.xlsx"
        )

    wb = Workbook()
    # 기본 빈 시트 제거
    wb.remove(wb.active)

    if selected_projects:
        # ── 프로젝트별 시트 세트 생성 ──────────────────────────
        for proj in selected_projects:
            df_proj = df[df["프로젝트"] == proj].copy()
            if df_proj.empty:
                logger.warning(f"프로젝트 '{proj}' 데이터 없음 — 시트 생략")
                continue

            stats_proj = generate_summary_stats(df_proj, report_month)
            team = PROJECT_TEAM_NAME_MAP.get(proj, proj.replace(" 업무 관리", "").strip())
            prefix = PROJECT_SHEET_PREFIX_MAP.get(proj, team[:4])

            _write_raw_data_sheet(wb, df_proj,
                                  sheet_name=f"{prefix}_정제데이터")
            _write_detail_sheet(wb, df_proj, stats_proj, team,
                                sheet_name=f"{prefix}_SR처리현황")
            _write_summary_sheet(wb, df_proj, stats_proj, team, report_month,
                                 sheet_name=f"{prefix}_SR집계")
            _write_dml_report_sheet(wb, stats_proj, team, report_month,
                                    sheet_name=f"{prefix}_DML처리현황")

        logger.info(f"{len(selected_projects)}개 프로젝트 시트 세트 생성 완료")
    else:
        # ── 단일 시트 세트 (기존 동작) ─────────────────────────
        _write_raw_data_sheet(wb, df)
        _write_detail_sheet(wb, df, stats, project_name)
        _write_summary_sheet(wb, df, stats, project_name, report_month)
        _write_dml_report_sheet(wb, stats, project_name, report_month)

    wb.save(output_path)
    logger.info(f"리포트 생성 완료: {output_path}")
    return output_path


# ─────────────────────────────────────────────────────────────
# 시트 1: 정제된 RAW 데이터
# ─────────────────────────────────────────────────────────────
def _write_raw_data_sheet(wb: Workbook, df: pd.DataFrame, sheet_name: str = "정제 데이터"):
    """정제된 RAW 데이터 시트 생성"""
    ws = wb.create_sheet(sheet_name)

    # 표시할 컬럼 순서
    display_cols = [
        "프로젝트", "키", "요약", "이슈 유형", "상태", "해결책",
        "담당자", "생성일", "변경일", "해결일",
        "시스템명", "시스템 부서", "서비스", "서비스등급",
        "업무구분(FTE)", "업무유형(FTE)",
        "처리상태", "문의유형",
        "접수일", "요청생성일",
        "처리시간_접수(일)", "처리시간_해결(일)",
    ]
    # 존재하는 컬럼만 필터
    cols = [c for c in display_cols if c in df.columns]

    # 헤더
    for col_idx, col_name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        _apply_cell_style(cell, FONT_HEADER_WHITE, HEADER_FILL_DARK, CENTER, THIN_BORDER)

    # 데이터
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, col_name in enumerate(cols, 1):
            value = row.get(col_name)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            _apply_cell_style(cell, FONT_NORMAL, border=THIN_BORDER, alignment=LEFT)

    # 컬럼 너비 자동 조절
    for col_idx in range(1, len(cols) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15

    # 자동 필터
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"
    ws.freeze_panes = "A2"


# ─────────────────────────────────────────────────────────────
# 시트 2: SR 처리 현황(상세)
# ─────────────────────────────────────────────────────────────
def _write_detail_sheet(wb: Workbook, df: pd.DataFrame, stats: dict, project_name: str,
                        sheet_name: str = "SR 처리 현황(상세)"):
    """SR 처리 현황(상세) 시트 - 시스템명 × 문의유형 × 처리상태 매트릭스"""
    ws = wb.create_sheet(sheet_name)

    # ── 타이틀 ──
    row = 1
    ws.cell(row=row, column=1, value=f"({project_name})")
    _apply_cell_style(ws.cell(row=row, column=1), FONT_TITLE)

    row = 2
    ws.cell(row=row, column=1, value="■ SR 처리 현황(상세)")
    _apply_cell_style(ws.cell(row=row, column=1), FONT_SECTION)

    ws.cell(row=row, column=20, value="(단위 : 건)")
    _apply_cell_style(ws.cell(row=row, column=20), FONT_NORMAL, alignment=RIGHT)

    # ── 헤더 구성 ──
    # Row 3: 구분 | (빈) | 요약 | 기능문의 ||| 단순조치/운영지원 ||| ...
    # Row 4: (빈) | 미접수 | 접수 | 완료 | 합계 | 미접수 | 접수 | 완료 | 종계 ...

    row = 3
    col = 1

    # 구분 헤더 (2행 병합)
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=1)
    cell = ws.cell(row=row, column=1, value="구분")
    _apply_cell_style(cell, FONT_HEADER_WHITE, HEADER_FILL_DARK, CENTER, THIN_BORDER)
    _apply_cell_style(ws.cell(row=row + 1, column=1), border=THIN_BORDER)

    col = 2

    # 요약 (미접수/접수/완료/합계) - 4컬럼
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 3)
    cell = ws.cell(row=row, column=col, value="요약")
    _apply_cell_style(cell, FONT_HEADER_WHITE, HEADER_FILL_DARK, CENTER, THIN_BORDER)
    for i, label in enumerate(["미접수", "접수", "완료", "합계"]):
        c = ws.cell(row=row + 1, column=col + i, value=label)
        _apply_cell_style(c, FONT_HEADER_BLACK, HEADER_FILL_LIGHT, CENTER, THIN_BORDER)
    col += 4

    # 각 문의유형별 4컬럼 (미접수, 접수, 완료, 종계)
    fills = [HEADER_FILL_LIGHT, HEADER_FILL_LIGHT, HEADER_FILL_LIGHT, HEADER_FILL_YELLOW, HEADER_FILL_YELLOW]
    for idx, itype in enumerate(INQUIRY_TYPE_ORDER):
        short_name = INQUIRY_SHORT_NAMES.get(itype, itype)
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 3)
        cell = ws.cell(row=row, column=col, value=short_name)
        fill = fills[idx] if idx < len(fills) else HEADER_FILL_LIGHT
        _apply_cell_style(cell, FONT_HEADER_BLACK, fill, CENTER, THIN_BORDER)
        for i, label in enumerate(["미접수", "접수", "완료", "종계"]):
            c = ws.cell(row=row + 1, column=col + i, value=label)
            _apply_cell_style(c, FONT_HEADER_BLACK, fill, CENTER, THIN_BORDER)
        col += 4

    # 빈 셀 border 처리
    for c in range(2, col):
        _apply_cell_style(ws.cell(row=row, column=c), border=THIN_BORDER)

    # ── 데이터 행 (시스템명별) ──
    row = 5
    systems = sorted(df["시스템명"].dropna().unique())

    for sys_name in systems:
        df_sys = df[df["시스템명"] == sys_name]

        ws.cell(row=row, column=1, value=sys_name)
        _apply_cell_style(ws.cell(row=row, column=1), FONT_NORMAL, alignment=LEFT, border=THIN_BORDER)

        col = 2
        # 요약: 미접수/접수/완료/합계
        status_counts = df_sys["처리상태"].value_counts()
        sys_total = len(df_sys)
        for status in PROCESS_STATUS_ORDER:
            val = int(status_counts.get(status, 0))
            c = ws.cell(row=row, column=col, value=val if val > 0 else "-")
            _apply_cell_style(c, FONT_NORMAL, alignment=CENTER, border=THIN_BORDER)
            col += 1
        c = ws.cell(row=row, column=col, value=sys_total)
        _apply_cell_style(c, FONT_NORMAL, alignment=CENTER, border=THIN_BORDER)
        col += 1

        # 각 문의유형별 미접수/접수/완료/종계
        for itype in INQUIRY_TYPE_ORDER:
            df_type = df_sys[df_sys["문의유형"] == itype]
            type_status = df_type["처리상태"].value_counts()
            type_total = len(df_type)
            for status in PROCESS_STATUS_ORDER:
                val = int(type_status.get(status, 0))
                c = ws.cell(row=row, column=col, value=val if val > 0 else "-")
                _apply_cell_style(c, FONT_NORMAL, alignment=CENTER, border=THIN_BORDER)
                col += 1
            c = ws.cell(row=row, column=col, value=type_total if type_total > 0 else "-")
            _apply_cell_style(c, FONT_NORMAL, alignment=CENTER, border=THIN_BORDER)
            col += 1

        row += 1

    # ── 총합계 행 ──
    ws.cell(row=row, column=1, value="총합계")
    _apply_cell_style(ws.cell(row=row, column=1), FONT_HEADER_BLACK, TOTAL_FILL, CENTER, THIN_BORDER)

    col = 2
    # 요약 총합계
    total_status = df["처리상태"].value_counts()
    for status in PROCESS_STATUS_ORDER:
        val = int(total_status.get(status, 0))
        c = ws.cell(row=row, column=col, value=val)
        _apply_cell_style(c, FONT_HEADER_BLACK, TOTAL_FILL, CENTER, THIN_BORDER)
        col += 1
    c = ws.cell(row=row, column=col, value=len(df))
    _apply_cell_style(c, FONT_HEADER_BLACK, TOTAL_FILL, CENTER, THIN_BORDER)
    col += 1

    # 문의유형별 총합계
    for itype in INQUIRY_TYPE_ORDER:
        df_type = df[df["문의유형"] == itype]
        type_status = df_type["처리상태"].value_counts()
        for status in PROCESS_STATUS_ORDER:
            val = int(type_status.get(status, 0))
            c = ws.cell(row=row, column=col, value=val)
            _apply_cell_style(c, FONT_HEADER_BLACK, TOTAL_FILL, CENTER, THIN_BORDER)
            col += 1
        c = ws.cell(row=row, column=col, value=len(df_type))
        _apply_cell_style(c, FONT_HEADER_BLACK, TOTAL_FILL, CENTER, THIN_BORDER)
        col += 1

    # 컬럼 너비
    ws.column_dimensions["A"].width = 28
    for c in range(2, col):
        ws.column_dimensions[get_column_letter(c)].width = 8

    ws.freeze_panes = "B5"


# ─────────────────────────────────────────────────────────────
# 시트 3: SR 건수 집계(요약) + 접수율/처리율 + 장기미처리 + 평균처리시간
# ─────────────────────────────────────────────────────────────
def _write_summary_sheet(
    wb: Workbook,
    df: pd.DataFrame,
    stats: dict,
    project_name: str,
    report_month: str,
    sheet_name: str = "SR 집계(요약)",
):
    """SR 건수 집계(요약) 및 부가 통계 시트"""
    ws = wb.create_sheet(sheet_name)

    row = 1

    # ── 타이틀 ──
    ws.cell(row=row, column=1, value=f"({project_name})(계속)")
    _apply_cell_style(ws.cell(row=row, column=1), FONT_TITLE)
    row += 2

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # ■ SR 건수 집계(요약)
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    ws.cell(row=row, column=1, value="■ SR 건수 집계(요약)")
    _apply_cell_style(ws.cell(row=row, column=1), FONT_SECTION)
    row += 1

    year = report_month.split("-")[0]
    ws.cell(row=row, column=1, value=f"※ {year}년 부터 생성된 SR 기준")
    _apply_cell_style(ws.cell(row=row, column=1), Font(name="맑은 고딕", size=9, italic=True))
    row += 1

    # 헤더
    headers_r1 = ["처리현황", "전체", "", "", "", "당월", "", "", ""]
    headers_r2 = ["", "미접수", "접수", "완료", "합계", "미접수", "접수", "완료", "합계"]

    # Row 1: 병합 헤더
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=1)
    cell = ws.cell(row=row, column=1, value="처리현황")
    _apply_cell_style(cell, FONT_HEADER_WHITE, HEADER_FILL_DARK, CENTER, THIN_BORDER)
    _apply_cell_style(ws.cell(row=row + 1, column=1), border=THIN_BORDER)

    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    cell = ws.cell(row=row, column=2, value="전체")
    _apply_cell_style(cell, FONT_HEADER_WHITE, HEADER_FILL_DARK, CENTER, THIN_BORDER)

    ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=9)
    cell = ws.cell(row=row, column=6, value="당월")
    _apply_cell_style(cell, FONT_HEADER_WHITE, HEADER_FILL_DARK, CENTER, THIN_BORDER)

    for c in range(2, 10):
        _apply_cell_style(ws.cell(row=row, column=c), border=THIN_BORDER)

    row += 1
    sub_headers = ["미접수", "접수", "완료", "합계", "미접수", "접수", "완료", "합계"]
    for i, h in enumerate(sub_headers, 2):
        cell = ws.cell(row=row, column=i, value=h)
        _apply_cell_style(cell, FONT_HEADER_BLACK, HEADER_FILL_LIGHT, CENTER, THIN_BORDER)
    row += 1

    # 건수 행
    ws.cell(row=row, column=1, value="건수")
    _apply_cell_style(ws.cell(row=row, column=1), FONT_HEADER_BLACK, HEADER_FILL_GRAY, CENTER, THIN_BORDER)
    total = stats["total"]
    monthly = stats["monthly"]
    values = [
        total["미접수"], total["접수"], total["완료"], total["합계"],
        monthly["미접수"], monthly["접수"], monthly["완료"], monthly["합계"],
    ]
    for i, v in enumerate(values, 2):
        cell = ws.cell(row=row, column=i, value=v)
        _apply_cell_style(cell, FONT_NORMAL, alignment=CENTER, border=THIN_BORDER)
    row += 1

    # % 행
    ws.cell(row=row, column=1, value="%")
    _apply_cell_style(ws.cell(row=row, column=1), FONT_HEADER_BLACK, HEADER_FILL_GRAY, CENTER, THIN_BORDER)
    for i in range(2, 10):
        _apply_cell_style(ws.cell(row=row, column=i), FONT_NORMAL, alignment=CENTER, border=THIN_BORDER)
    row += 1

    # 접수율/처리율 텍스트
    rates = stats["rates"]
    ws.cell(row=row, column=1, value=f"• 접수율({rates['접수율']}%) (접수+완료)/전체")
    row += 1
    ws.cell(row=row, column=1, value=f"• 처리율({rates['처리율']}%) 완료/전")
    row += 2

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # ■ 접수율 및 처리율
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    ws.cell(row=row, column=11, value="■ 접수율 및 처리율")
    _apply_cell_style(ws.cell(row=row, column=11), FONT_SECTION)
    row_rates_start = row
    row += 1

    # 헤더
    rate_headers = ["문의유형", "기능문의", "단순조치/\n운영지원", "데이터\n추출/수정", "신규\n개발/개선", "운영관리", "합계"]
    for i, h in enumerate(rate_headers):
        col = 11 + i
        cell = ws.cell(row=row, column=col, value=h)
        _apply_cell_style(cell, FONT_HEADER_WHITE, HEADER_FILL_DARK, CENTER, THIN_BORDER)
    row += 1

    # 건수 행
    ws.cell(row=row, column=11, value="건수")
    _apply_cell_style(ws.cell(row=row, column=11), FONT_HEADER_BLACK, HEADER_FILL_GRAY, CENTER, THIN_BORDER)
    inquiry_rates = stats["inquiry_rates"]
    total_count = 0
    for i, itype in enumerate(INQUIRY_TYPE_ORDER):
        cnt = inquiry_rates.get(itype, {}).get("건수", 0)
        total_count += cnt
        cell = ws.cell(row=row, column=12 + i, value=cnt)
        _apply_cell_style(cell, FONT_NORMAL, alignment=CENTER, border=THIN_BORDER)
    cell = ws.cell(row=row, column=17, value=total_count)
    _apply_cell_style(cell, FONT_NORMAL, alignment=CENTER, border=THIN_BORDER)
    row += 1

    # % 행
    ws.cell(row=row, column=11, value="%")
    _apply_cell_style(ws.cell(row=row, column=11), FONT_HEADER_BLACK, HEADER_FILL_GRAY, CENTER, THIN_BORDER)
    for i, itype in enumerate(INQUIRY_TYPE_ORDER):
        cnt = inquiry_rates.get(itype, {}).get("건수", 0)
        pct = round(cnt / total_count * 100, 1) if total_count > 0 else 0
        cell = ws.cell(row=row, column=12 + i, value=f"{pct}%")
        _apply_cell_style(cell, FONT_NORMAL, alignment=CENTER, border=THIN_BORDER)
    cell = ws.cell(row=row, column=17, value="100%")
    _apply_cell_style(cell, FONT_NORMAL, alignment=CENTER, border=THIN_BORDER)
    row += 2

    # 분류 설명
    descriptions = [
        "• 기능문의 : 시스템/업무 사용법 질의, 정책 근거 해석요청, 답변안내 등",
        "• 단순조치/운영지원 : 계정/권한, 월마감, 단순처리/전달, 일반 지원 등",
        "• 데이터 추출/수정 : 데이터 추출, 변경/이관/적재 등",
        "• 신규 개발 및 개선 : 신규 기능 개발, 기능변경, 오류수정",
        "• 운영관리 : S/W/DB작업, 품질/장애예방, 보안, 라이선스 관리, 교육 등",
    ]
    for desc in descriptions:
        ws.cell(row=row, column=11, value=desc)
        _apply_cell_style(ws.cell(row=row, column=11), Font(name="맑은 고딕", size=8))
        row += 1

    row += 1

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # ■ 장기 미처리 SR
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    lp_row = row
    ws.cell(row=lp_row, column=1, value="■ 장기 미처리 SR (건수)")
    _apply_cell_style(ws.cell(row=lp_row, column=1), FONT_SECTION)
    lp_row += 1

    month_parts = report_month.split("-")
    ws.cell(
        row=lp_row, column=1,
        value=f"• 기준일({month_parts[0]}/{month_parts[1]}말) 기준, -90일 이전 접수된 미처리 SR 건수"
    )
    _apply_cell_style(ws.cell(row=lp_row, column=1), Font(name="맑은 고딕", size=9))
    lp_row += 1

    # 장기 미처리 테이블
    lp_headers = ["구분", "미접수", "접수", "합계", "주요 미처리 사유"]
    for i, h in enumerate(lp_headers):
        cell = ws.cell(row=lp_row, column=1 + i, value=h)
        _apply_cell_style(cell, FONT_HEADER_WHITE, HEADER_FILL_DARK, CENTER, THIN_BORDER)
    lp_row += 1

    long_pending = stats.get("long_pending", pd.DataFrame())
    if len(long_pending) > 0:
        lp_status = long_pending["처리상태"].value_counts()
        lp_pending = int(lp_status.get("미접수", 0))
        lp_accepted = int(lp_status.get("접수", 0))
        lp_total = lp_pending + lp_accepted
    else:
        lp_pending, lp_accepted, lp_total = 0, 0, 0

    ws.cell(row=lp_row, column=1, value="시스템 또는 서비스\n레벨")
    _apply_cell_style(ws.cell(row=lp_row, column=1), FONT_NORMAL, alignment=CENTER, border=THIN_BORDER)
    for i, v in enumerate([lp_pending, lp_accepted, lp_total]):
        cell = ws.cell(row=lp_row, column=2 + i, value=v if v > 0 else "-")
        _apply_cell_style(cell, FONT_NORMAL, alignment=CENTER, border=THIN_BORDER)
    ws.cell(row=lp_row, column=5, value="")
    _apply_cell_style(ws.cell(row=lp_row, column=5), FONT_NORMAL, alignment=LEFT, border=THIN_BORDER)

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # ■ 평균처리 시간(일)
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    avg_row = lp_row - 3  # 장기 미처리와 같은 높이에 배치
    ws.cell(row=avg_row, column=11, value="■ 평균처리 시간(일)")
    _apply_cell_style(ws.cell(row=avg_row, column=11), FONT_SECTION)
    avg_row += 2

    avg_headers = ["당월", "평균 처리시간(일)"]
    for i, h in enumerate(avg_headers):
        cell = ws.cell(row=avg_row, column=11 + i, value=h)
        _apply_cell_style(cell, FONT_HEADER_WHITE, HEADER_FILL_DARK, CENTER, THIN_BORDER)
    avg_row += 1

    avg_times = stats.get("avg_times_monthly", {})
    avg_data = [
        ("생성일 → 접수일", avg_times.get("생성일_접수일", "-")),
        ("생성일 → 완료일", avg_times.get("생성일_완료일", "-")),
    ]
    for label, val in avg_data:
        ws.cell(row=avg_row, column=11, value=label)
        _apply_cell_style(ws.cell(row=avg_row, column=11), FONT_NORMAL, alignment=LEFT, border=THIN_BORDER)
        display_val = val if val is not None else "-"
        ws.cell(row=avg_row, column=12, value=display_val)
        _apply_cell_style(ws.cell(row=avg_row, column=12), FONT_NORMAL, alignment=CENTER, border=THIN_BORDER)
        avg_row += 1

    # 컬럼 너비 설정
    ws.column_dimensions["A"].width = 16
    for c in range(2, 10):
        ws.column_dimensions[get_column_letter(c)].width = 10
    ws.column_dimensions["K"].width = 16
    for c in range(12, 18):
        ws.column_dimensions[get_column_letter(c)].width = 12


# ─────────────────────────────────────────────────────────────
# 시트 4: DML 처리 현황 (3. 데이터 추출 및 수정)
# ─────────────────────────────────────────────────────────────
def _write_dml_report_sheet(
    wb: Workbook,
    stats: dict,
    project_name: str,
    report_month: str,
    sheet_name: str = "DML 처리 현황",
):
    """DML(데이터 추출 및 수정) 처리 현황 시트 — 시스템별 당월 집계 및 주요 사유"""
    ws = wb.create_sheet(sheet_name)
    dml_data = stats.get("data_request_report", [])

    year, month = report_month.split("-")
    month_label = f"{int(month)}월"

    # ── 타이틀 ──
    row = 1
    ws.cell(row=row, column=1, value=f"({project_name})(계속)")
    _apply_cell_style(ws.cell(row=row, column=1), FONT_TITLE)
    row += 1

    ws.cell(row=row, column=1, value=f"■ DML 처리 현황 — {month_label}")
    _apply_cell_style(ws.cell(row=row, column=1), FONT_SECTION)
    ws.cell(row=row, column=7, value="(단위 : 건)")
    _apply_cell_style(ws.cell(row=row, column=7), FONT_NORMAL, alignment=RIGHT)
    row += 1

    # ── 헤더 2행 ──
    # 구분 (rowspan 2)
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=1)
    c = ws.cell(row=row, column=1, value="구분")
    _apply_cell_style(c, FONT_HEADER_WHITE, HEADER_FILL_DARK, CENTER, THIN_BORDER)
    _apply_cell_style(ws.cell(row=row + 1, column=1), border=THIN_BORDER)

    # 당월 (colspan 4)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    c = ws.cell(row=row, column=2, value=month_label)
    _apply_cell_style(c, FONT_HEADER_WHITE, HEADER_FILL_DARK, CENTER, THIN_BORDER)
    for i in range(3, 6):
        _apply_cell_style(ws.cell(row=row, column=i), border=THIN_BORDER)

    # 주요 사유 (rowspan 2)
    ws.merge_cells(start_row=row, start_column=6, end_row=row + 1, end_column=6)
    c = ws.cell(row=row, column=6, value="주요 사유")
    _apply_cell_style(c, FONT_HEADER_WHITE, HEADER_FILL_DARK, CENTER, THIN_BORDER)
    _apply_cell_style(ws.cell(row=row + 1, column=6), border=THIN_BORDER)

    # 향후 대책 (rowspan 2)
    ws.merge_cells(start_row=row, start_column=7, end_row=row + 1, end_column=7)
    c = ws.cell(row=row, column=7, value="향후 대책")
    _apply_cell_style(c, FONT_HEADER_WHITE, HEADER_FILL_DARK, CENTER, THIN_BORDER)
    _apply_cell_style(ws.cell(row=row + 1, column=7), border=THIN_BORDER)

    row += 1
    # 서브헤더: 미접수 접수 완료 합계
    for i, label in enumerate(["미접수", "접수", "완료", "합계"], 2):
        c = ws.cell(row=row, column=i, value=label)
        _apply_cell_style(c, FONT_HEADER_BLACK, HEADER_FILL_LIGHT, CENTER, THIN_BORDER)
    row += 1

    # ── 데이터 행 ──
    total_미접수 = total_접수 = total_완료 = total_합계 = 0

    for r in dml_data:
        ws.cell(row=row, column=1, value=r["시스템명"])
        _apply_cell_style(ws.cell(row=row, column=1), FONT_NORMAL, alignment=LEFT, border=THIN_BORDER)

        for col_idx, key in enumerate(["미접수", "접수", "완료", "합계"], 2):
            val = r[key]
            c = ws.cell(row=row, column=col_idx, value=val if val > 0 else "-")
            _apply_cell_style(c, FONT_NORMAL, alignment=CENTER, border=THIN_BORDER)

        total_미접수 += r["미접수"]
        total_접수 += r["접수"]
        total_완료 += r["완료"]
        total_합계 += r["합계"]

        # 주요 사유 (wrap_text 포함)
        c = ws.cell(row=row, column=6, value=r["주요_사유"])
        _apply_cell_style(c, FONT_NORMAL, alignment=LEFT, border=THIN_BORDER)

        # 향후 대책 (빈 칸)
        c = ws.cell(row=row, column=7, value="")
        _apply_cell_style(c, FONT_NORMAL, alignment=LEFT, border=THIN_BORDER)

        # 행 높이: 주요 사유 줄 수에 맞게 조절 (1줄 = 15pt)
        n_lines = max(1, len(r["주요_사유_목록"]))
        ws.row_dimensions[row].height = max(20, n_lines * 15)
        row += 1

    # ── 총합계 ──
    ws.cell(row=row, column=1, value="총합계")
    _apply_cell_style(ws.cell(row=row, column=1), FONT_HEADER_BLACK, TOTAL_FILL, CENTER, THIN_BORDER)
    for col_idx, val in enumerate([total_미접수, total_접수, total_완료, total_합계], 2):
        c = ws.cell(row=row, column=col_idx, value=val)
        _apply_cell_style(c, FONT_HEADER_BLACK, TOTAL_FILL, CENTER, THIN_BORDER)
    for col_idx in range(6, 8):
        c = ws.cell(row=row, column=col_idx, value="")
        _apply_cell_style(c, FONT_HEADER_BLACK, TOTAL_FILL, CENTER, THIN_BORDER)

    # ── 컬럼 너비 ──
    ws.column_dimensions["A"].width = 30
    for col_idx in range(2, 6):
        ws.column_dimensions[get_column_letter(col_idx)].width = 9
    ws.column_dimensions["F"].width = 70
    ws.column_dimensions["G"].width = 25

    ws.freeze_panes = "B5"
